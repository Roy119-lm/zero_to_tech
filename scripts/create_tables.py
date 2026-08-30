import pandas as pd
from openpyxl import load_workbook
from pyspark.sql import SparkSession

WAREHOUSE = "s3://processed-data/warehouse"

file_path = "/opt/spark/data/indicator_system.xlsx"
target_sheet = "20251121"
header_idx = 2

df = pd.read_excel(
    file_path,
    sheet_name=target_sheet,
    header=header_idx,
    engine="openpyxl",
)
wb = load_workbook(file_path, data_only=True)
ws = wb[target_sheet]

color_indicator_type_map = {"FFFFC000": "政务指标", "00000000": "运营商指标", "FF00B0F0": "遥感指标", "FF92D050": "电力指标"}
indicator_type = []
color_column_idx = 4
for row_num in range(header_idx + 2, ws.max_row + 1):
    cell = ws.cell(row=row_num, column=color_column_idx)
    color_hex = str(cell.fill.start_color.index)
    indicator_type.append(color_indicator_type_map[color_hex])

df["indicator_type"] = indicator_type

indicator_cols = ["一级指标", "二级指标", "三级", "四级"]
df[indicator_cols] = df[indicator_cols].ffill()

# ========== 四级指标表定义 ==========
dataset_table = {
    "ts": "timestamp",
    "region_code": "int",   # 行政区划代码
    "province": "string",
    "city": "string",
    "county": "string",
    "year": "int",
    "season": "int",
    "indicator_name": "string",
    "indicator_value": "decimal(20, 8)",
    "extra_info": "string", # 额外信息，例如具体指标的地域划分不是完全按照行政区划分割时可以将其具体的位置信息存储在这一项
}

# 创建SparkSession，启用Iceberg
spark = (
    SparkSession.builder.appName("Iceberg Native IO")
    .getOrCreate()
)

spark.sql("CREATE DATABASE IF NOT EXISTS indicators")

# 筛选出元数据列，丢弃不需要的列
meta_columns = [
    "一级指标",
    "二级指标",
    "三级",
    "四级",
    "数据类型",
    "口径",
    "数据来源",
    "indicator_type",   # 指标类型：政务指标、运营商指标、遥感指标、电力指标
]

# 选取数据并重命名为英文列名（Iceberg 推荐英文列名）
dict_df = df[meta_columns].copy()
print(dict_df)
dict_df.columns = [
    "l1_name",
    "l2_name",
    "l3_name",
    "l4_name",
    "data_type",
    "indicator_equation",
    "indicator_source",
    "indicator_type",
]

spark_dict_df = spark.createDataFrame(dict_df)
# 写入 Iceberg 表
print("正在写入指标字典表...")
spark_dict_df.writeTo(
    "indicators.dictionary"
).createOrReplace()

col_defs = []
for col_name, col_type in dataset_table.items():
    spark_type = col_type
    col_defs.append(f"{col_name} {spark_type}")

col_str = ", ".join(col_defs)

ddl = f"""
CREATE TABLE IF NOT EXISTS indicators.data (
    {col_str}
)
USING iceberg
PARTITIONED BY (days(ts))
"""

# 3. 执行
print(f"正在创建指标数据表...")
spark.sql(ddl)

print("所有业务指标表结构创建完毕！")
